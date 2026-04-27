from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse
from typing import List, Dict, Any
import fitz
import pandas as pd
import uuid
import os
import re
import json

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from openai import OpenAI

app = FastAPI()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None


# ---------------- KNOWLEDGE BASE FALLBACK ----------------

WORK_KB = {
    "cable": {
        "keywords": ["cable", "xlpe", "ht cable", "lt cable", "armoured", "unarmoured", "termination", "glanding"],
        "work": "Cable Laying / Cable Termination Work",
        "material": "Cable, cable lugs, glands, ferrules, cable tags, saddles/clamps, cable ties, warning tape if specified.",
        "method": "Check route and drawings, place cable drum on jack, use rollers, pull cable safely, dress cable, gland and terminate, test insulation and continuity.",
        "tools": "Cable roller, drum jack, winch, crimping tool, megger, multimeter, torque tools, PPE.",
        "labour": "Cable gang 4-6 helpers, electrician/technician and site supervisor.",
        "time": "Approx. 300-500 meter/day for normal route depending on cable size, route and manpower."
    },
    "earthing": {
        "keywords": ["earthing", "earth pit", "earth electrode", "gi strip", "copper strip", "earth mat", "chemical earthing"],
        "work": "Earthing Work",
        "material": "Earth electrode, GI/Copper strip, earth compound/charcoal/salt as specified, chamber cover and connection accessories.",
        "method": "Excavate pit, install electrode, connect strip, backfill with compound, watering and earth resistance testing.",
        "tools": "Earth tester, spade, welding/drilling tools, multimeter, PPE.",
        "labour": "Electrician, 2-3 labour and supervisor.",
        "time": "Approx. 2-4 hours per earth pit depending on soil and site condition."
    },
    "concrete": {
        "keywords": ["concrete", "rcc", "pcc", "m10", "m15", "m20", "m25", "m30", "cement concrete"],
        "work": "Concrete Work",
        "material": "Cement, sand, aggregate, water, admixture if specified and reinforcement steel if RCC.",
        "method": "Batching, mixing, placing, compaction/vibration, finishing and curing as per tender specification.",
        "tools": "Concrete mixer, vibrator, trowel, level, cube mould, curing arrangement.",
        "labour": "Mason, 4-6 labour and supervisor.",
        "time": "Approx. 8-12 Cum/day with small mixer team, depending on site condition."
    },
    "excavation": {
        "keywords": ["excavation", "earthwork", "earth work", "trench", "pit", "digging", "backfilling", "soil"],
        "work": "Excavation / Earthwork",
        "material": "No permanent material; backfilling material/disposal as per tender requirement.",
        "method": "Marking, excavation manually or by machine, dressing, dewatering if required, disposal/backfilling and compaction.",
        "tools": "JCB, spade, pickaxe, dumper, measuring tape, level instrument.",
        "labour": "Machine operator or 4-6 labour with supervisor.",
        "time": "Depends on depth, soil type and method. Machine excavation is faster than manual."
    },
    "demolition": {
        "keywords": ["demolition", "dismantling", "dismantle", "breaking", "removal", "dispose", "disposal"],
        "work": "Demolition / Dismantling Work",
        "material": "No permanent material required; debris handling and consumables only.",
        "method": "Site inspection, utility isolation, barricading, controlled dismantling/breaking, stacking of serviceable material and disposal of debris.",
        "tools": "Breaker, hammer, chisel, cutter, wheelbarrow, barricade, PPE.",
        "labour": "Demolition labour 4-6 persons and supervisor.",
        "time": "Depends on structure size, access and safety restriction."
    },
    "panel": {
        "keywords": ["panel", "switchgear", "mccb", "acb", "lt panel", "control panel", "distribution board", "db"],
        "work": "Electrical Panel Work",
        "material": "Panel, breaker, busbar, control wiring, lugs, glands, ferrules and panel accessories.",
        "method": "Position panel, align and fix, cable termination, control wiring, testing and commissioning.",
        "tools": "Crimping tool, multimeter, torque wrench, megger, screwdriver set.",
        "labour": "Electrician, technician and supervisor.",
        "time": "Approx. 1-2 days per panel depending on size and termination quantity."
    },
    "scada": {
        "keywords": ["scada", "rtu", "plc", "hmi", "automation", "remote monitoring", "data logger", "communication"],
        "work": "SCADA / Automation Work",
        "material": "PLC/RTU, HMI, sensors, communication module, control cable, panel accessories.",
        "method": "Install devices, complete wiring, configure communication, test signals and commission system.",
        "tools": "Laptop, multimeter, crimping tool, communication tester.",
        "labour": "Automation engineer and technician.",
        "time": "Approx. 1-3 days depending on points and communication complexity."
    },
    "painting": {
        "keywords": ["painting", "paint", "primer", "coating", "enamel", "epoxy", "white wash", "distemper"],
        "work": "Painting / Coating Work",
        "material": "Primer, paint/coating, thinner, putty if specified.",
        "method": "Surface preparation, primer coat, paint/coating application and finishing.",
        "tools": "Brush, roller, spray gun, scraper, sandpaper.",
        "labour": "Painter and helper.",
        "time": "Depends on area, surface condition and number of coats."
    },
    "masonry": {
        "keywords": ["brick", "brickwork", "block work", "masonry", "aac block", "stone masonry"],
        "work": "Masonry Work",
        "material": "Bricks/blocks/stone, cement mortar, sand and water.",
        "method": "Line-level setting, mortar preparation, laying, joint filling and curing.",
        "tools": "Trowel, line dori, level tube, plumb bob, scaffolding.",
        "labour": "Mason and helper.",
        "time": "Depends on wall thickness, height and site access."
    },
    "plaster": {
        "keywords": ["plaster", "plastering", "cement plaster"],
        "work": "Plaster Work",
        "material": "Cement, sand, water and curing arrangement.",
        "method": "Surface preparation, mortar mixing, plaster application, levelling, finishing and curing.",
        "tools": "Trowel, level, wooden float, scaffolding.",
        "labour": "Mason and helper.",
        "time": "Depends on area and thickness."
    },
    "solar": {
        "keywords": ["solar", "module", "pv", "inverter", "string", "mms", "dc cable", "ac cable"],
        "work": "Solar Plant Work",
        "material": "PV module, MMS, inverter, cables, connectors, earthing and accessories as specified.",
        "method": "Install structure/modules, do cabling, termination, testing and commissioning.",
        "tools": "Torque wrench, multimeter, megger, MC4 crimper, PPE.",
        "labour": "Solar technician, electrician and supervisor.",
        "time": "Depends on plant size and manpower."
    }
}


# ---------------- PDF EXTRACTION ----------------

def extract_pdf_lines(file_bytes: bytes) -> List[Dict[str, Any]]:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    rows = []

    for page_no, page in enumerate(doc, start=1):
        text = page.get_text("text")
        for line in text.splitlines():
            clean = re.sub(r"\s+", " ", line).strip()
            if clean:
                rows.append({
                    "page": page_no,
                    "text": clean
                })

    return rows


# ---------------- SECTION DETECTION ----------------

def detect_section(text: str) -> str:
    t = text.lower()

    if any(x in t for x in ["bill of quantity", "boq", "schedule of quantity", "price schedule", "schedule-b"]):
        return "BOQ"
    if any(x in t for x in ["scope of work", "scope"]):
        return "SCOPE_OF_WORK"
    if any(x in t for x in ["technical specification", "technical specifications", "specification", "specifications"]):
        return "TECHNICAL_SPECIFICATION"
    if any(x in t for x in ["approved make", "make list", "list of approved makes", "approved brand"]):
        return "APPROVED_MAKE"
    if any(x in t for x in ["vendor list", "approved vendor", "vendors"]):
        return "VENDOR_LIST"
    if any(x in t for x in ["general condition", "gcc", "general terms"]):
        return "GENERAL_CONDITION"
    if any(x in t for x in ["special condition", "scc", "special terms"]):
        return "SPECIAL_CONDITION"
    if any(x in t for x in ["safety", "ppe", "hse", "work permit"]):
        return "SAFETY"
    if any(x in t for x in ["drawing", "drawing no", "drawing number", "drg"]):
        return "DRAWING"
    if any(x in t for x in ["standard", "is code", "iec", "ieee", "bis", "cpwd", "morth"]):
        return "STANDARD"

    return "OTHER"


def build_context(lines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    current_section = "OTHER"
    output = []

    for row in lines:
        section = detect_section(row["text"])
        if section != "OTHER":
            current_section = section

        output.append({
            "page": row["page"],
            "section": current_section,
            "text": row["text"]
        })

    return output


# ---------------- BOQ EXTRACTION ----------------

def normalize_unit(unit: str) -> str:
    if not unit:
        return "-"
    u = unit.strip().lower()

    mapping = {
        "cum": "Cum", "cu.m": "Cum", "cu.m.": "Cum", "m3": "Cum", "m³": "Cum", "cmt": "Cum",
        "sqm": "Sqm", "sq.m": "Sqm", "sq.m.": "Sqm", "m2": "Sqm", "m²": "Sqm", "smt": "Sqm",
        "rmt": "Rmt", "rm": "Rmt", "mtr": "Mtr", "meter": "Mtr", "metre": "Mtr", "m": "Mtr",
        "nos": "Nos", "no": "Nos", "each": "Nos", "set": "Set", "job": "Job", "lot": "Lot",
        "kg": "Kg", "mt": "MT", "ton": "MT", "ltr": "Ltr", "litre": "Ltr"
    }

    return mapping.get(u, unit)


def merge_boq_lines(context_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged = []
    buffer = ""
    page = 1
    section = "OTHER"

    item_start = re.compile(r"^\s*(\d+(\.\d+)*|[A-Z]\d+)\s+")

    for row in context_rows:
        line = row["text"]

        if item_start.match(line):
            if buffer:
                merged.append({"page": page, "section": section, "text": buffer.strip()})
            buffer = line
            page = row["page"]
            section = row["section"]
        else:
            if buffer:
                buffer += " " + line

    if buffer:
        merged.append({"page": page, "section": section, "text": buffer.strip()})

    return merged


def extract_boq(context_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    boq = []
    rows = merge_boq_lines(context_rows)

    units = r"(cum|cu\.m\.?|m3|m³|cmt|sqm|sq\.m\.?|m2|m²|smt|rmt|rm|mtr|meter|metre|m|nos|no|each|set|job|lot|kg|mt|ton|ltr|litre)"

    pattern = re.compile(
        rf"^\s*(?P<item>\d+(\.\d+)*|[A-Z]\d+)\s+"
        rf"(?P<desc>.*?)\s+"
        rf"(?P<qty>\d+(\.\d+)?)\s*"
        rf"(?P<unit>{units})\s+"
        rf"(?P<rate>\d+(\.\d+)?)",
        re.IGNORECASE
    )

    for row in rows:
        clean = row["text"].replace(",", "").replace("₹", "")
        m = pattern.search(clean)

        if m:
            boq.append({
                "Item No": m.group("item"),
                "Description": m.group("desc").strip(),
                "Quantity": m.group("qty"),
                "Unit": normalize_unit(m.group("unit")),
                "Rate": m.group("rate"),
                "Page": row["page"],
                "Section": row["section"],
                "Source Line": row["text"]
            })

    return boq


# ---------------- CLAUSE / CONTEXT MAPPING ----------------

def find_global_section_text(context_rows: List[Dict[str, Any]], section_name: str, limit: int = 50) -> str:
    selected = [r for r in context_rows if r["section"] == section_name]
    return "\n".join([f"[Page {r['page']}] {r['text']}" for r in selected[:limit]])


def get_relevant_context(description: str, context_rows: List[Dict[str, Any]], max_lines: int = 45) -> str:
    words = set(re.findall(r"[a-zA-Z0-9]{4,}", description.lower()))
    scored = []

    important_sections = [
        "SCOPE_OF_WORK",
        "TECHNICAL_SPECIFICATION",
        "APPROVED_MAKE",
        "VENDOR_LIST",
        "SPECIAL_CONDITION",
        "GENERAL_CONDITION",
        "SAFETY",
        "DRAWING",
        "STANDARD"
    ]

    for row in context_rows:
        text_lower = row["text"].lower()
        score = 0

        for w in words:
            if w in text_lower:
                score += 3

        if row["section"] in important_sections:
            score += 2

        if any(x in text_lower for x in ["shall", "must", "as per", "is ", "iec", "approved", "make", "vendor", "drawing", "safety"]):
            score += 1

        if score > 0:
            scored.append((score, row))

    scored.sort(reverse=True, key=lambda x: x[0])
    top = scored[:max_lines]

    return "\n".join([f"[Page {r['page']} | {r['section']}] {r['text']}" for _, r in top])


def extract_rule_based_make_vendor(context_text: str) -> Dict[str, str]:
    lower = context_text.lower()
    approved_make = "-"
    vendor_list = "-"
    standards = "-"

    make_lines = []
    vendor_lines = []
    standard_lines = []

    for line in context_text.splitlines():
        l = line.lower()

        if any(x in l for x in ["approved make", "make:", "make -", "brand", "manufacturer"]):
            make_lines.append(line)

        if any(x in l for x in ["vendor", "supplier", "authorized dealer", "approved vendor"]):
            vendor_lines.append(line)

        if any(x in l for x in ["is ", "iec", "ieee", "bis", "cpwd", "morth", "standard", "as per"]):
            standard_lines.append(line)

    if make_lines:
        approved_make = "\n".join(make_lines[:8])

    if vendor_lines:
        vendor_list = "\n".join(vendor_lines[:8])

    if standard_lines:
        standards = "\n".join(standard_lines[:8])

    return {
        "approved_make": approved_make,
        "vendor_list": vendor_list,
        "standards": standards
    }


# ---------------- FALLBACK ANALYSIS ----------------

def fallback_analysis(description: str, unit: str, context_text: str) -> Dict[str, str]:
    desc = description.lower()
    unit_l = str(unit).lower()

    best = None
    best_score = 0

    for _, data in WORK_KB.items():
        score = 0
        for kw in data["keywords"]:
            if kw in desc:
                score += 3

        if unit_l in ["cum"] and any(x in desc for x in ["concrete", "rcc", "pcc", "excavation", "earth"]):
            score += 2
        if unit_l in ["sqm"] and any(x in desc for x in ["painting", "plaster", "floor", "tile"]):
            score += 2
        if unit_l in ["mtr", "rmt"] and any(x in desc for x in ["cable", "pipe", "strip"]):
            score += 2

        if score > best_score:
            best_score = score
            best = data

    rule_extract = extract_rule_based_make_vendor(context_text)

    if best:
        confidence = "High" if best_score >= 5 else "Medium"
        return {
            "content_type": "WORK_ITEM",
            "identified_work_type": best["work"],
            "related_conditions": "-",
            "technical_specification": "-",
            "applicable_standard": rule_extract["standards"],
            "material_required": best["material"],
            "approved_make": rule_extract["approved_make"],
            "vendor_list": rule_extract["vendor_list"],
            "execution_method": best["method"],
            "tools_equipment": best["tools"],
            "labour_requirement": best["labour"],
            "supervisor_requirement": "Site supervisor / engineer required for quality and safety control.",
            "estimated_time": best["time"],
            "drawing_reference": "-",
            "risk_review_point": "-",
            "confidence": confidence
        }

    return {
        "content_type": "WORK_ITEM",
        "identified_work_type": "General Tender Work Item",
        "related_conditions": "-",
        "technical_specification": "-",
        "applicable_standard": rule_extract["standards"],
        "material_required": "As per tender item description, specification and approved make list if applicable.",
        "approved_make": rule_extract["approved_make"],
        "vendor_list": rule_extract["vendor_list"],
        "execution_method": "Execute as per tender specification, drawings, applicable standards and site engineer instruction.",
        "tools_equipment": "Standard tools, tackles, measuring instruments and PPE as required.",
        "labour_requirement": "Skilled / semi-skilled labour as required.",
        "supervisor_requirement": "Site supervisor required.",
        "estimated_time": "To be estimated based on quantity, site access, manpower and method.",
        "drawing_reference": "-",
        "risk_review_point": "Work type not strongly identified. Verify with tender specification.",
        "confidence": "Low"
    }


# ---------------- AI JSON ----------------

def safe_json_loads(text: str):
    try:
        clean = text.strip()
        clean = clean.replace("```json", "").replace("```", "").strip()
        return json.loads(clean)
    except Exception:
        return None


def ai_analyze_item(item: Dict[str, Any], relevant_context: str, global_make: str, global_vendor: str) -> Dict[str, str]:
    fallback = fallback_analysis(item["Description"], item["Unit"], relevant_context + "\n" + global_make + "\n" + global_vendor)

    if client is None:
        fallback["risk_review_point"] = "AI disabled because OPENAI_API_KEY not found. Fallback used."
        return fallback

    prompt = f"""
You are a senior EPC tender analysis engineer.

You must deeply understand tender documents.

TASK:
For the BOQ item below, identify:
1. Whether it is a work item, condition, specification or other.
2. What exact work is required.
3. Related tender conditions/clauses.
4. Technical specifications.
5. Applicable standards like IS, IEC, IEEE, CPWD, MORTH, BIS etc.
6. Materials required.
7. Approved make / brand list if tender provides.
8. Vendor list if tender provides.
9. Method of execution.
10. Tools and equipment.
11. Labour and supervisor requirement.
12. Practical estimated execution time.
13. Drawing reference if available.
14. Risk/review point.
15. Confidence.

IMPORTANT RULES:
- Do NOT hallucinate tender-specific approved makes/vendors/standards.
- If tender context does not provide approved make/vendor/standard, write "-".
- Method, tools, labour and time may use engineering knowledge if tender is silent.
- Always return valid JSON only.
- confidence must be High, Medium or Low.

RETURN JSON ONLY with these keys:
{{
  "content_type": "",
  "identified_work_type": "",
  "related_conditions": "",
  "technical_specification": "",
  "applicable_standard": "",
  "material_required": "",
  "approved_make": "",
  "vendor_list": "",
  "execution_method": "",
  "tools_equipment": "",
  "labour_requirement": "",
  "supervisor_requirement": "",
  "estimated_time": "",
  "drawing_reference": "",
  "risk_review_point": "",
  "confidence": ""
}}

BOQ ITEM:
Item No: {item["Item No"]}
Description: {item["Description"]}
Quantity: {item["Quantity"]}
Unit: {item["Unit"]}
Rate: {item["Rate"]}

RELEVANT TENDER CONTEXT:
{relevant_context}

GLOBAL APPROVED MAKE SECTION:
{global_make}

GLOBAL VENDOR SECTION:
{global_vendor}
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a professional EPC tender analyst. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0
        )

        content = response.choices[0].message.content
        data = safe_json_loads(content)

        if not data:
            fallback["risk_review_point"] = "AI response JSON parsing failed. Fallback used."
            return fallback

        final = fallback.copy()
        for key in final:
            value = data.get(key)
            if value is not None and str(value).strip():
                final[key] = str(value).strip()

        return final

    except Exception as e:
        fallback["risk_review_point"] = f"AI failed. Fallback used. Error: {str(e)}"
        fallback["confidence"] = "Medium" if fallback["confidence"] == "High" else fallback["confidence"]
        return fallback


# ---------------- EXCEL FORMAT ----------------

def get_float(value):
    try:
        return float(str(value).replace(",", "").replace("₹", ""))
    except Exception:
        return 0.0


def format_excel(file_name: str):
    wb = load_workbook(file_name)
    ws = wb.active
    ws.title = "Vtenders Pro Output"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    header_fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin

    ws.freeze_panes = "A2"

    confidence_col = None
    risk_col = None

    for cell in ws[1]:
        if cell.value == "Confidence":
            confidence_col = cell.column
        if cell.value == "Risk / Review Point":
            risk_col = cell.column

    for row in range(2, ws.max_row + 1):
        if confidence_col:
            c = ws.cell(row=row, column=confidence_col)
            if str(c.value).lower() == "high":
                c.fill = high_fill
            elif str(c.value).lower() == "medium":
                c.fill = medium_fill
            else:
                c.fill = low_fill

        if risk_col:
            r = ws.cell(row=row, column=risk_col)
            if r.value and str(r.value).strip() not in ["-", ""]:
                r.fill = medium_fill

    widths = {
        "A": 10, "B": 58, "C": 30, "D": 12, "E": 10, "F": 12, "G": 16,
        "H": 18, "I": 55, "J": 60, "K": 55, "L": 45, "M": 55, "N": 45,
        "O": 45, "P": 55, "Q": 45, "R": 40, "S": 35, "T": 35, "U": 45,
        "V": 14, "W": 25, "X": 80
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 60

    wb.save(file_name)


# ---------------- ROUTES ----------------

@app.get("/", response_class=HTMLResponse)
def home():
    ai_status = "Enabled" if client else "Disabled - OPENAI_API_KEY missing"
    return f"""
    <html>
    <body style="font-family: Arial; padding: 30px;">
        <h2>Vtenders Pro AI Tender Analyzer</h2>
        <p>AI Status: <b>{ai_status}</b></p>
        <p>Upload tender PDF files. Output includes work identification, clause mapping, specifications, approved make, vendor list, method, tools, labour and time.</p>
        <form action="/upload/" method="post" enctype="multipart/form-data">
            <input type="file" name="files" multiple required>
            <br><br>
            <button type="submit">Upload & Analyze</button>
        </form>
    </body>
    </html>
    """


@app.get("/ai-test")
def ai_test():
    sample_item = {
        "Item No": "1",
        "Description": "Providing and laying 3.5 core 240 sqmm XLPE armoured cable including termination",
        "Quantity": "100",
        "Unit": "Mtr",
        "Rate": "500",
        "Page": 1,
        "Section": "BOQ",
        "Source Line": "-"
    }

    result = ai_analyze_item(
        sample_item,
        "Cable shall be laid as per approved route drawing. Cable shall conform to relevant IS/IEC standard.",
        "Approved make: Polycab, Havells, KEI.",
        "Approved vendor: Authorized dealer only."
    )

    return {
        "ai_enabled": client is not None,
        "result": result
    }


@app.post("/upload/")
async def upload(files: List[UploadFile] = File(...)):
    output = []

    for upload_file in files:
        content = await upload_file.read()

        lines = extract_pdf_lines(content)
        context_rows = build_context(lines)

        global_make = find_global_section_text(context_rows, "APPROVED_MAKE", limit=80)
        global_vendor = find_global_section_text(context_rows, "VENDOR_LIST", limit=80)

        boq_items = extract_boq(context_rows)

        for item in boq_items:
            relevant_context = get_relevant_context(item["Description"], context_rows, max_lines=45)
            analysis = ai_analyze_item(item, relevant_context, global_make, global_vendor)

            qty = get_float(item["Quantity"])
            rate = get_float(item["Rate"])
            amount = qty * rate

            output.append({
                "Item No": item["Item No"],
                "Tender Work Item": item["Description"],
                "Identified Work Type": analysis["identified_work_type"],
                "Quantity": qty,
                "Unit": item["Unit"],
                "Rate": rate,
                "Amount": amount,
                "Content Type": analysis["content_type"],
                "Related Tender Conditions": analysis["related_conditions"],
                "Technical Specification": analysis["technical_specification"],
                "Applicable Standard": analysis["applicable_standard"],
                "Material Required": analysis["material_required"],
                "Approved Make": analysis["approved_make"],
                "Vendor List": analysis["vendor_list"],
                "Execution Method": analysis["execution_method"],
                "Tools & Equipment": analysis["tools_equipment"],
                "Labour Requirement": analysis["labour_requirement"],
                "Supervisor Requirement": analysis["supervisor_requirement"],
                "Estimated Time": analysis["estimated_time"],
                "Drawing Reference": analysis["drawing_reference"],
                "Risk / Review Point": analysis["risk_review_point"],
                "Confidence": analysis["confidence"],
                "Source Page / Section": f"Page {item['Page']} / {item['Section']}",
                "Source Line": item["Source Line"]
            })

    if not output:
        output.append({
            "Item No": "-",
            "Tender Work Item": "No BOQ Data Found",
            "Identified Work Type": "PDF format not supported or BOQ not detected",
            "Quantity": "-",
            "Unit": "-",
            "Rate": "-",
            "Amount": "-",
            "Content Type": "REVIEW",
            "Related Tender Conditions": "-",
            "Technical Specification": "-",
            "Applicable Standard": "-",
            "Material Required": "-",
            "Approved Make": "-",
            "Vendor List": "-",
            "Execution Method": "-",
            "Tools & Equipment": "-",
            "Labour Requirement": "-",
            "Supervisor Requirement": "-",
            "Estimated Time": "-",
            "Drawing Reference": "-",
            "Risk / Review Point": "No BOQ rows detected. PDF may be scanned/image based or table layout unsupported.",
            "Confidence": "Low",
            "Source Page / Section": "-",
            "Source Line": "-"
        })

    df = pd.DataFrame(output)

    file_name = f"vtenders_pro_output_{uuid.uuid4().hex}.xlsx"
    df.to_excel(file_name, index=False)

    format_excel(file_name)

    return FileResponse(
        path=file_name,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/download/{file_name}")
def download_file(file_name: str):
    file_path = os.path.join(os.getcwd(), file_name)
    return FileResponse(path=file_path, filename=file_name)
